from __future__ import annotations

import math
import os
import re
import sys
import threading
from collections import Counter
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
try:
    from pypinyin import lazy_pinyin
except ImportError:  # Development fallback; the Windows build always bundles pypinyin.
    lazy_pinyin = None


APP_NAME = "机票航段整理工具"
MONTHS = {m: i for i, m in enumerate(
    ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"], 1
)}
ROUTE_RE = re.compile(r"\b[A-Z]{3}(?:\s*-\s*[A-Z]{3})+\b")
DATE_RE = re.compile(r"(?<![A-Z0-9])(\d{1,2})\s*([A-Z]{3})(?![A-Z])", re.I)
TICKET_RE = re.compile(r"(?:\bAIR\s*TICKET|\bAITICKET|\bTICKET\b)", re.I)
AIRPORT_ALIASES = {
    "成都天府": "TFU", "天府": "TFU", "成都天府国际机场": "TFU",
    "亚的斯亚贝巴": "ADD", "亚的斯亚贝巴博莱": "ADD", "亚的斯亚贝巴博莱国际机场": "ADD",
    "维多利亚瀑布": "VFA", "维多利亚瀑布机场": "VFA",
    "广州白云": "CAN", "广州白云国际机场": "CAN",
    "北京首都": "PEK", "上海浦东": "PVG", "哈拉雷": "HRE", "布拉瓦约": "BUQ",
    "卢萨卡": "LUN", "开普敦": "CPT", "马普托": "MPM",
}

# This small fallback covers the user's historical workbook. The packaged EXE also
# contains the complete airportsdata IATA database and works without the internet.
FALLBACK_AIRPORTS = {
    "ADD": ("Addis Ababa Bole International Airport", "HAAB", 8.9779, 38.7993, "Ethiopia"),
    "BUQ": ("Joshua Mqabuko Nkomo International Airport", "FVBU", -20.0174, 28.6179, "Zimbabwe"),
    "CAN": ("Guangzhou Baiyun International Airport", "ZGGG", 23.3924, 113.2988, "China"),
    "CPT": ("Cape Town International Airport", "FACT", -33.9697, 18.5972, "South Africa"),
    "HRE": ("Robert Gabriel Mugabe International Airport", "FVRG", -17.9318, 31.0928, "Zimbabwe"),
    "LUN": ("Kenneth Kaunda International Airport", "FLKK", -15.3308, 28.4526, "Zambia"),
    "MPM": ("Maputo International Airport", "FQMA", -25.9208, 32.5726, "Mozambique"),
    "PEK": ("Beijing Capital International Airport", "ZBAA", 40.0801, 116.5846, "China"),
    "PVG": ("Shanghai Pudong International Airport", "ZSPD", 31.1434, 121.8052, "China"),
    "TFU": ("Chengdu Tianfu International Airport", "ZUTF", 30.3125, 104.441, "China"),
    "VFA": ("Victoria Falls International Airport", "FVFA", -18.0959, 25.8390, "Zimbabwe"),
}


def load_airports():
    result = dict(FALLBACK_AIRPORTS)
    try:
        import airportsdata
        for code, item in airportsdata.load("IATA").items():
            lat, lon = item.get("lat"), item.get("lon")
            if lat is None or lon is None:
                continue
            result[code.upper()] = (
                item.get("name", ""), item.get("icao", ""), float(lat), float(lon), item.get("country", "")
            )
    except Exception:
        pass
    return result


AIRPORTS = load_airports()


def airport_code(label):
    value = normalize(label).replace("国际机场", "").replace("机场", "").strip()
    if re.fullmatch(r"[A-Za-z]{3}", value):
        return value.upper()
    return AIRPORT_ALIASES.get(value, "")


def chinese_to_pinyin(name):
    """Convert a Chinese name to uppercase Hanyu Pinyin, separated by spaces."""
    if not name or not re.search(r"[\u3400-\u9fff]", str(name)):
        return ""
    if lazy_pinyin is None:
        return str(name).strip()
    return " ".join(lazy_pinyin(str(name).strip())).upper()


def haversine(lat1, lon1, lat2, lon2):
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return round(2 * r * math.asin(math.sqrt(a)), 2)


def normalize(text):
    return re.sub(r"\s+", " ", str(text).replace("–", "-").replace("—", "-")).strip()


def extract_dates(text, year):
    upper = text.upper()
    # If a change instruction supplies replacement dates, use the final change section.
    change = re.search(r"(?:CHANGE(?:D)?\s+(?:DATE(?:S)?\s*)?(?:TO\s*)?)(.+)$", upper)
    date_source = change.group(1) if change and "CHANGED FROM" not in upper else upper
    dates = []
    for day, mon in DATE_RE.findall(date_source):
        try:
            dates.append(datetime(year, MONTHS[mon.upper()], int(day)))
        except ValueError:
            continue
    # Handle compact ranges such as 16-23APR or 11-14MAY.
    for d1, d2, mon in re.findall(r"(?<!\d)(\d{1,2})\s*-\s*(\d{1,2})\s*([A-Z]{3})", date_source, re.I):
        try:
            pair = [datetime(year, MONTHS[mon.upper()], int(d1)), datetime(year, MONTHS[mon.upper()], int(d2))]
            if not dates:
                dates.extend(pair)
            else:
                for d in pair:
                    if d not in dates:
                        dates.append(d)
        except ValueError:
            pass
    return dates


def parse_record(raw, source_row, year):
    text = normalize(raw)
    upper = text.upper()
    routes = ROUTE_RE.findall(upper)
    if not routes:
        return [], "未识别到机场路线"
    first_route = re.search(ROUTE_RE, upper)
    ticket = TICKET_RE.search(upper)
    name = ""
    if ticket and first_route:
        middle = text[ticket.end():first_route.start()].strip(" -")
        # Some records place a date before the route; remove it from the passenger name.
        middle = DATE_RE.sub("", middle).strip(" -")
        middle = re.sub(r"\bCHANGE(?:D)?\b.*$", "", middle, flags=re.I).strip(" -")
        name = middle
    dates = extract_dates(text, year)
    cancelled = any(k in upper for k in ("CANCELLED", "CANCELLATION FEE", "REFUND", "REFUNDED"))
    notes = []
    if cancelled: notes.append("取消/退款")
    if "CHANGE" in upper: notes.append("改签")
    if "BUSINESS" in upper: notes.append("商务舱")
    if "NO DRT" in upper: notes.append("不含接送")
    if "INC DRT" in upper: notes.append("含接送")

    legs = []
    leg_index = 0
    for route_text in routes:
        codes = [x.strip() for x in route_text.split("-")]
        for a, b in zip(codes, codes[1:]):
            leg_index += 1
            dt = dates[min(leg_index - 1, len(dates) - 1)] if dates else None
            warning = []
            if not name: warning.append("姓名未识别")
            if dt is None: warning.append("日期缺失")
            if a == b: warning.append("起降机场相同")
            if a not in AIRPORTS: warning.append(f"未知机场:{a}")
            if b not in AIRPORTS: warning.append(f"未知机场:{b}")
            dist = None
            if a in AIRPORTS and b in AIRPORTS:
                dist = haversine(AIRPORTS[a][2], AIRPORTS[a][3], AIRPORTS[b][2], AIRPORTS[b][3])
            legs.append({"source_row": source_row, "leg": leg_index, "name": name, "chinese_name": "",
                         "english_name": name, "date": dt,
                         "from": a, "to": b, "distance": dist, "notes": "; ".join(notes),
                         "include": "否" if cancelled else "是", "warning": "; ".join(warning), "raw": text})
    return legs, ""


def parse_full_datetime(value):
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M", "%Y-%m-%d", "%Y/%m/%d"):
        try: return datetime.strptime(s, fmt)
        except ValueError: pass
    return None


def extract_records(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    records = []
    # A previously generated workbook contains the same text in several output
    # sheets. Prefer its dedicated raw-data sheet to avoid counting it twice.
    sheets = [wb["原始记录"]] if "原始记录" in wb.sheetnames else [
        ws for ws in wb.worksheets if ws.title not in {"航段明细", "机场坐标", "异常复核", "说明"}
    ]
    # First support travel-agent statements with separate named columns.
    for ws in sheets:
        header_row = None
        header_map = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 30), values_only=True), 1):
            labels = [str(v).strip() if v is not None else "" for v in row]
            if "中文行程" in labels and "航班日期" in labels:
                header_row = row_idx
                header_map = {v: i for i, v in enumerate(labels) if v}
                break
        if header_row:
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
                def val(name):
                    idx = header_map.get(name)
                    return row[idx] if idx is not None and idx < len(row) else None
                itinerary = val("中文行程")
                if not itinerary: continue
                labels = [normalize(x) for x in re.split(r"\s*-\s*", str(itinerary)) if normalize(x)]
                codes = [airport_code(x) for x in labels]
                dates = [parse_full_datetime(x) for x in str(val("航班日期") or "").split("/")]
                dates = [x for x in dates if x]
                eng_name = re.sub(r"\s+(MR|MRS|MS|MISS)$", "", str(val("乘客姓名") or "").strip(), flags=re.I)
                cn_name = str(val("中文姓名") or "").strip()
                status = str(val("订单状态") or "").strip()
                dept = str(val("部门") or "").strip()
                raw = " | ".join(str(x) for x in row if x not in (None, ""))
                structured = {"labels": labels, "codes": codes, "dates": dates,
                              "name": cn_name or eng_name, "english_name": eng_name,
                              "status": status, "department": dept}
                records.append({"sheet": ws.title, "row": row_idx, "raw": raw, "structured": structured})
            if records:
                return records
    for ws in sheets:
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            for value in row:
                if isinstance(value, str) and TICKET_RE.search(value) and ROUTE_RE.search(value.upper()):
                    records.append({"sheet": ws.title, "row": row_idx, "raw": normalize(value), "structured": None})
                    break
    return records


def parse_structured_record(rec, source_row):
    data = rec["structured"]
    codes, labels, dates = data["codes"], data["labels"], data["dates"]
    cancelled = any(k in data["status"] for k in ("取消", "退款", "退票"))
    legs = []
    for idx, (a, b) in enumerate(zip(codes, codes[1:]), 1):
        warning = []
        if not a: warning.append(f"未知机场:{labels[idx-1]}")
        if not b: warning.append(f"未知机场:{labels[idx]}")
        dt = dates[min(idx - 1, len(dates) - 1)] if dates else None
        if dt is None: warning.append("日期缺失")
        dist = None
        if a in AIRPORTS and b in AIRPORTS:
            dist = haversine(AIRPORTS[a][2], AIRPORTS[a][3], AIRPORTS[b][2], AIRPORTS[b][3])
        notes = "; ".join(x for x in [data["department"], data["status"], data["english_name"]] if x)
        english_name = data["english_name"] or chinese_to_pinyin(data["name"])
        legs.append({"source_row": source_row, "leg": idx, "name": data["name"],
                     "chinese_name": data["name"] if re.search(r"[\u3400-\u9fff]", data["name"]) else "",
                     "english_name": english_name, "date": dt,
                     "from": a or labels[idx-1], "to": b or labels[idx], "distance": dist,
                     "notes": notes, "include": "否" if cancelled else "是",
                     "warning": "; ".join(warning), "raw": rec["raw"]})
    return legs


def style_sheet(ws, widths):
    fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.fill, c.font, c.alignment = fill, Font(color="FFFFFF", bold=True), Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def process_file(input_path, output_path, year):
    records = extract_records(input_path)
    all_legs, failed = [], []
    for seq, rec in enumerate(records, 1):
        if rec["structured"]:
            legs, error = parse_structured_record(rec, seq), ""
        else:
            legs, error = parse_record(rec["raw"], seq, year)
        if error: failed.append((seq, rec["sheet"], rec["row"], error, rec["raw"]))
        all_legs.extend(legs)
    dupes = Counter(x["raw"].upper() for x in records)

    wb = Workbook()
    ws = wb.active
    ws.title = "航段明细"
    headers = ["原始序号", "来源工作表", "来源行号", "航段序号", "姓名", "中文姓名", "英文姓名", "日期", "出发机场", "出发纬度",
               "出发经度", "到达机场", "到达纬度", "到达经度", "大圆距离(km)", "备注", "建议计入后续核算",
               "异常/复核提示", "疑似重复原始记录", "原始记录"]
    ws.append(headers)
    for leg in all_legs:
        src = records[leg["source_row"] - 1]
        fa, ta = AIRPORTS.get(leg["from"]), AIRPORTS.get(leg["to"])
        ws.append([leg["source_row"], src["sheet"], src["row"], leg["leg"], leg["name"],
                   leg["chinese_name"], leg["english_name"], leg["date"], leg["from"],
                   fa[2] if fa else None, fa[3] if fa else None, leg["to"], ta[2] if ta else None,
                   ta[3] if ta else None, leg["distance"], leg["notes"], leg["include"], leg["warning"],
                   "是" if dupes[src["raw"].upper()] > 1 else "", leg["raw"]])
    for cell in ws["H"][1:]: cell.number_format = "yyyy-mm-dd hh:mm"
    style_sheet(ws, [10, 14, 10, 10, 20, 14, 24, 18, 11, 12, 12, 11, 12, 12, 16, 20, 18, 26, 18, 70])

    used = sorted({leg[x] for leg in all_legs for x in ("from", "to")})
    ap = wb.create_sheet("机场坐标")
    ap.append(["IATA", "机场名称", "ICAO", "纬度", "经度", "国家/地区", "数据来源"])
    for code in used:
        a = AIRPORTS.get(code)
        ap.append([code, *(a or ("", "", None, None, "")), "airportsdata (OpenFlights/OurAirports)"])
    style_sheet(ap, [10, 42, 10, 13, 13, 18, 34])

    raw_ws = wb.create_sheet("原始记录")
    raw_ws.append(["原始序号", "来源工作表", "来源行号", "原始票务记录"])
    for seq, rec in enumerate(records, 1): raw_ws.append([seq, rec["sheet"], rec["row"], rec["raw"]])
    style_sheet(raw_ws, [12, 18, 12, 90])

    review = wb.create_sheet("异常复核")
    review.append(["原始序号", "来源工作表", "来源行号", "问题", "原始记录"])
    for item in failed: review.append(item)
    for leg in all_legs:
        if leg["warning"]:
            src = records[leg["source_row"] - 1]
            review.append([leg["source_row"], src["sheet"], src["row"], leg["warning"], leg["raw"]])
    style_sheet(review, [12, 18, 12, 30, 90])

    info = wb.create_sheet("说明")
    info.append(["项目", "内容"])
    info.append(["程序", APP_NAME])
    info.append(["生成时间", datetime.now()])
    info.append(["行程年份", year])
    info.append(["处理结果", f"识别原始记录 {len(records)} 条，拆分航段 {len(all_legs)} 条，异常复核 {review.max_row-1} 条。"])
    info.append(["距离算法", "Haversine 大圆距离，地球平均半径 6371 km；结果不含碳排放量。"])
    info.append(["数据安全", "全程在本机离线处理，不上传原始表格。"])
    info.append(["姓名转换", "已有英文姓名时原样保留；只有中文姓名时转换为大写汉语拼音。拼音不等同于护照拼写，正式申报前请核对。"])
    style_sheet(info, [20, 95])
    wb.save(output_path)
    return len(records), len(all_legs), review.max_row - 1


class App:
    def __init__(self, root):
        self.root = root
        root.title(APP_NAME)
        root.geometry("720x330")
        root.resizable(False, False)
        self.path = tk.StringVar()
        self.year = tk.StringVar(value=str(datetime.now().year))
        self.status = tk.StringVar(value="请选择包含票务记录的 Excel 文件。")
        frame = ttk.Frame(root, padding=24)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text=APP_NAME, font=("Microsoft YaHei UI", 18, "bold")).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 24))
        ttk.Label(frame, text="原始表格：").grid(row=1, column=0, sticky="e", pady=8)
        ttk.Entry(frame, textvariable=self.path, width=65).grid(row=1, column=1, sticky="ew", pady=8)
        ttk.Button(frame, text="选择文件", command=self.choose).grid(row=1, column=2, padx=(10, 0))
        ttk.Label(frame, text="行程年份：").grid(row=2, column=0, sticky="e", pady=8)
        ttk.Spinbox(frame, from_=2000, to=2100, textvariable=self.year, width=10).grid(row=2, column=1, sticky="w", pady=8)
        ttk.Button(frame, text="开始整理", command=self.start, width=20).grid(row=3, column=1, sticky="w", pady=24)
        ttk.Separator(frame).grid(row=4, column=0, columnspan=3, sticky="ew")
        ttk.Label(frame, textvariable=self.status, foreground="#1F4E78", wraplength=650).grid(row=5, column=0, columnspan=3, sticky="w", pady=16)
        frame.columnconfigure(1, weight=1)

    def choose(self):
        p = filedialog.askopenfilename(filetypes=[("Excel 工作簿", "*.xlsx *.xlsm")])
        if p: self.path.set(p)

    def start(self):
        p = Path(self.path.get())
        if not p.exists():
            messagebox.showwarning(APP_NAME, "请先选择有效的 Excel 文件。")
            return
        try: year = int(self.year.get())
        except ValueError:
            messagebox.showwarning(APP_NAME, "行程年份必须是四位数字。")
            return
        out = p.with_name(p.stem + "_航段整理结果.xlsx")
        self.status.set("正在处理，请稍候……")
        def work():
            try:
                stats = process_file(p, out, year)
                self.root.after(0, lambda: self.done(out, stats))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror(APP_NAME, f"处理失败：\n{e}"))
                self.root.after(0, lambda: self.status.set("处理失败，请检查文件格式。"))
        threading.Thread(target=work, daemon=True).start()

    def done(self, out, stats):
        self.status.set(f"完成：原始记录 {stats[0]} 条，航段 {stats[1]} 条，需复核 {stats[2]} 条。\n输出：{out}")
        messagebox.showinfo(APP_NAME, f"处理完成！\n\n输出文件：\n{out}")
        try: os.startfile(out.parent)
        except Exception: pass


def main():
    if len(sys.argv) >= 3 and sys.argv[1] == "--test":
        stats = process_file(Path(sys.argv[2]), Path("test_output.xlsx"), int(sys.argv[3]) if len(sys.argv) > 3 else 2026)
        print(f"records={stats[0]} legs={stats[1]} review={stats[2]}")
        return
    root = tk.Tk()
    try: root.iconname(APP_NAME)
    except Exception: pass
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
